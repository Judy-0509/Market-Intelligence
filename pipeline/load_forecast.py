"""연전망(Smartphone SET TAM) 엑셀 -> SQLite 적재 CLI.

사용법:
    python load_forecast.py --source "<연전망.xlsx>" --sheet "6월 연전망용" --vintage 2026-06 [--db tam.db]

소스 파일 구조 (B3:BA18 범위, 사내 EDM 다운로드 기준):
  - 4행 = 헤더 (Vendor + 분기 라벨 + 연간 라벨)
  - 5~18행 = 벤더별 데이터 (14개 벤더, B열에 벤더명)
  - '25/'26/'27 연간 컬럼만 사용, 분기 컬럼은 무시

콘솔 출력(print/에러 메시지)은 전부 영문 ASCII로만 작성한다 (사내 콘솔이 cp949일 수 있어
출력 중 인코딩 에러로 죽으면 안 되기 때문). 한글은 docstring/주석에만 사용한다.
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import tam_db

HEADER_ROW = 4
FIRST_VENDOR_ROW = 5
LAST_VENDOR_ROW = 18
MIN_COL = 2  # column B
MAX_COL = 53  # column BA
NEEDED_YEARS = (2025, 2026, 2027)

# 확정된 캐노니컬 벤더명 (14개). 소스 파일의 원문 표기는 다를 수 있어 정규화 후 매칭하되,
# DB에는 항상 이 표기로 저장한다.
CANONICAL_VENDORS = [
    "MX",
    "Apple",
    "Huawei",
    "Honor",
    "Oppo",
    "Vivo",
    "Xiaomi",
    "Lenovo",
    "Transsion",
    "CN others",
    "Google",
    "HMD (Nokia)",
    "Local",
    "Total",
]

VINTAGE_RE = re.compile(r"^(\d{4})-(\d{2})$")


class PipelineError(Exception):
    """소스 파일 파싱/검증 실패 시 사용하는 예외. 메시지는 사용자에게 그대로 출력된다."""


def _safe_print(text: str, *, file=None) -> None:
    """print()가 cp949 등 제한된 콘솔 코드페이지에서 UnicodeEncodeError로 죽지 않도록 감싼다."""
    stream = file if file is not None else sys.stdout
    try:
        print(text, file=stream)
    except UnicodeEncodeError:
        encoding = getattr(stream, "encoding", None) or "ascii"
        print(text.encode(encoding, errors="replace").decode(encoding), file=stream)


# ---------------------------------------------------------------------------
# locate_year_columns() -- CONFIRMED against the in-house update_tam.py source
# on 2026-07-07. This is the one isolated function to keep in sync if
# update_tam.py's labeling rule ever changes again.
# ---------------------------------------------------------------------------
def locate_year_columns(header_cells) -> dict:
    """헤더 행에서 '25/'26/'27 연간 컬럼의 위치를 찾는다.

    header_cells: (column_index, raw_cell_value) 튜플의 iterable (헤더 행, B~BA열).
    반환: {year: column_index} (2025/2026/2027 각각 정확히 1개 컬럼).

    CONFIRMED 규칙 (update_tam.py, 2026-07-07):
    연간 라벨은 앞에 어퍼스트로피(')가 붙은 문자열 리터럴이다 (예: "'25", "'26", "'27").
    이 어퍼스트로피가 분기 라벨("1Q26", "1Q18" 등, 어퍼스트로피 없음)과의 구분 기준이다.
    이 파일들은 기계적으로 생성되므로 어퍼스트로피는 셀 텍스트에 저장된 리터럴 문자이며
    엑셀의 "텍스트로 강제" 접두 마커가 아니다. 문자열이 아닌 셀 값(숫자/날짜)은 연간
    라벨이 될 수 없다.

    update_tam.py와 의도적으로 다른 점: 같은 연도에 라벨이 중복되면 update_tam.py는 마지막
    컬럼을 조용히 채택하지만, 우리는 잘못 채택 시 vintage 전체가 오염되므로 즉시 에러를 낸다.
    """
    matches: dict = {year: [] for year in NEEDED_YEARS}
    scanned = []
    token_to_year = {"25": 2025, "26": 2026, "27": 2027}
    for col_idx, raw_value in header_cells:
        col_letter = get_column_letter(col_idx)
        scanned.append((col_letter, raw_value))
        if not isinstance(raw_value, str):
            continue
        text = raw_value.strip()
        if not text.startswith("'"):
            continue
        token = text.lstrip("'")
        year = token_to_year.get(token)
        if year is not None:
            matches[year].append(col_idx)

    missing = [year for year in NEEDED_YEARS if not matches[year]]
    ambiguous = {year: cols for year, cols in matches.items() if len(cols) > 1}
    if missing or ambiguous:
        listing = "\n".join(f"  {letter}: {raw!r}" for letter, raw in scanned)
        raise PipelineError(
            "Cannot locate annual forecast year columns unambiguously.\n"
            f"Missing years: {missing}\n"
            f"Ambiguous years (column indexes): {ambiguous}\n"
            f"Header row {HEADER_ROW} scanned (column: raw label):\n{listing}"
        )
    return {year: matches[year][0] for year in NEEDED_YEARS}


def _normalize_vendor(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip()).casefold()


def extract_vendor_names(vendor_rows) -> list:
    """벤더 데이터 행에서 (캐노니컬 벤더명, 원본 row) 쌍의 리스트를 반환한다.

    B열이 빈 칸인 행은 건너뛴다 (update_tam.py와 동일). 단, 건너뛴 뒤에도 14개 캐노니컬
    벤더명이 전부 존재해야 하며, 그렇지 않거나 알 수 없는 이름이 나오면 즉시 에러를 낸다.
    """
    normalized_lookup = {_normalize_vendor(name): name for name in CANONICAL_VENDORS}
    raw_names_seen = []
    pairs = []
    for row in vendor_rows:
        cell = row[0]  # column B
        raw = cell.value
        raw_str = "" if raw is None else str(raw).strip()
        if raw_str == "":
            continue  # blank vendor cell: skip (mirrors update_tam.py)
        raw_names_seen.append(raw_str)
        canonical = normalized_lookup.get(_normalize_vendor(raw_str))
        if canonical is None:
            raise PipelineError(
                "Unknown vendor name in source file.\n"
                f"  Cell {cell.coordinate}: {raw_str!r}\n"
                f"  Expected vendors: {CANONICAL_VENDORS}\n"
                f"  Found vendors: {raw_names_seen}"
            )
        pairs.append((canonical, row))

    found_names = [canonical for canonical, _ in pairs]
    missing = [v for v in CANONICAL_VENDORS if v not in found_names]
    if missing:
        raise PipelineError(
            "Missing expected vendor(s) in source file.\n"
            f"  Missing: {missing}\n"
            f"  Expected vendors: {CANONICAL_VENDORS}\n"
            f"  Found vendors: {found_names}"
        )
    return pairs


def parse_numeric_cell(cell) -> "float | None":
    """셀 값을 숫자(float) 또는 None으로 변환한다.

    빈 칸(None 또는 공백 문자열)은 None으로 저장한다 (0으로 임의 치환하지 않음 --
    사내 tam.xlsx 업데이터는 자기 목적상 0으로 치환하지만, 이 DB는 원본 그대로 저장한다).
    숫자가 아닌 비어있지 않은 값은 셀 위치와 함께 즉시 에러를 낸다.
    """
    value = cell.value
    if value is None:
        return None
    if isinstance(value, bool):
        raise PipelineError(f"Non-numeric value at {cell.coordinate}: {value!r}")
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.strip() == "":
        return None
    raise PipelineError(f"Non-numeric value at {cell.coordinate}: {value!r}")


def extract_forecast(source_path, sheet_name) -> list:
    """소스 엑셀에서 (vendor, year, value) 튜플 리스트를 추출한다."""
    source_path = Path(source_path)
    if not source_path.exists():
        raise PipelineError(f"Source file not found: {source_path}")

    wb = load_workbook(source_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise PipelineError(
                f"Sheet {sheet_name!r} not found. Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        table_rows = list(
            ws.iter_rows(min_row=HEADER_ROW, max_row=LAST_VENDOR_ROW, min_col=MIN_COL, max_col=MAX_COL)
        )
    finally:
        wb.close()

    header_row = table_rows[0]
    vendor_rows = table_rows[1:]
    expected_vendor_rows = LAST_VENDOR_ROW - FIRST_VENDOR_ROW + 1
    if len(vendor_rows) != expected_vendor_rows:
        raise PipelineError(
            f"Expected {expected_vendor_rows} vendor rows ({FIRST_VENDOR_ROW}-{LAST_VENDOR_ROW}), "
            f"got {len(vendor_rows)}"
        )

    # NOTE: cells past the last populated column in a row come back as openpyxl's
    # read-only EmptyCell, which has no .column/.coordinate attribute (only .value).
    # Column indexes are derived positionally from MIN_COL instead of cell.column.
    header_cells = [(MIN_COL + i, cell.value) for i, cell in enumerate(header_row)]
    year_cols = locate_year_columns(header_cells)

    vendor_pairs = extract_vendor_names(vendor_rows)

    rows = []
    for vendor_name, row in vendor_pairs:
        row_by_col = {MIN_COL + i: cell for i, cell in enumerate(row)}
        for year in NEEDED_YEARS:
            cell = row_by_col[year_cols[year]]
            value = parse_numeric_cell(cell)
            rows.append((vendor_name, year, value))
    return rows


def validate_vintage(vintage: str) -> None:
    m = VINTAGE_RE.match(vintage)
    if not m:
        raise PipelineError(f"Invalid --vintage format: {vintage!r} (expected YYYY-MM)")
    month = int(m.group(2))
    if not (1 <= month <= 12):
        raise PipelineError(f"Invalid --vintage month: {vintage!r} (month must be 01-12)")


def parse_args(argv=None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load a monthly forecast Excel file into the TAM SQLite DB."
    )
    parser.add_argument("--source", required=True, help="Path to the forecast .xlsx file")
    parser.add_argument("--sheet", required=True, help="Sheet name (Korean sheet name from the source workbook)")
    parser.add_argument("--vintage", required=True, help="Forecast vintage, format YYYY-MM")
    parser.add_argument("--db", default="tam.db", help="SQLite DB path (default: tam.db)")
    return parser.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)

    try:
        validate_vintage(args.vintage)
        rows = extract_forecast(args.source, args.sheet)
        conn = tam_db.init_db(Path(args.db))
        try:
            tam_db.upsert_vintage(conn, args.vintage, rows, source_file=str(args.source))
        finally:
            conn.close()
    except PipelineError as exc:
        _safe_print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:  # unexpected failure -- still fail loudly, never crash silently
        _safe_print(f"ERROR: unexpected failure: {exc}", file=sys.stderr)
        return 1

    vendor_count = len({vendor for vendor, _, _ in rows})
    years = sorted({year for _, year, _ in rows})
    total_2026 = next((value for vendor, year, value in rows if vendor == "Total" and year == 2026), None)
    _safe_print(f"Loaded vintage {args.vintage}: {vendor_count} vendors, years {years}")
    _safe_print(f"Sanity check - Total/2026: {total_2026}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
