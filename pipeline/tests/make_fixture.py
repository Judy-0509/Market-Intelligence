"""연전망(Smartphone SET TAM) 유사 구조의 합성(synthetic) 엑셀 픽스처 생성기.

실제 사내 파일이 아닌 테스트용 가짜 데이터를 만든다 (.gitignore가 *.xlsx를 이미
차단하므로 커밋되지 않는다). 헤더 라벨은 update_tam.py로 확정된 실제 소스 파일 형태를
흉내낸다: 연간 라벨은 어퍼스트로피 접두 문자열("'25", "'26", "'27" 등), 분기 라벨은
어퍼스트로피 없는 문자열("1Q18", "1Q26" 등).
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook

from load_forecast import CANONICAL_VENDORS  # noqa: E402

HEADER_ROW = 4
FIRST_VENDOR_ROW = 5
LAST_VENDOR_ROW = 18
VENDOR_COL = 2  # column B

# 헤더 레이아웃: (컬럼 인덱스, 라벨). update_tam.py 확정 규칙을 반영한 실제 형태 흉내.
BASE_HEADERS = [
    (2, "Vendor"),
    (3, "'17"),
    (4, "1Q18"),
    (5, "2Q18"),
    (6, "'25"),
    (7, "1Q26"),
    (8, "'26"),
    (9, "'27"),
    (10, "'31"),
]
YEAR_COL = {2025: 6, 2026: 8, 2027: 9}


def vendor_value(vendor_index: int, year: int, month_variant: int) -> float:
    """벤더 순번/연도/월(6 또는 7)에 따른 결정론적 예측값 (테스트 기대값의 근거)."""
    base = 100 * (vendor_index + 1) + (year - 2025) * 5
    if month_variant == 7:
        base += vendor_index  # 7월 소폭 리비전
    return float(base)


def build_workbook(
    sheet_name: str,
    headers=None,
    vendors=None,
    month_variant: int = 6,
    value_overrides=None,
    blank_cells=None,
) -> Workbook:
    """단일 시트 워크북을 만든다.

    headers: (컬럼, 라벨) 리스트. 기본은 BASE_HEADERS.
    vendors: 5~18행에 채울 벤더명 리스트 (14개). 기본은 CANONICAL_VENDORS.
    value_overrides: {(vendor, year): value} -- 특정 셀 값을 강제 지정 (리비전/NULL-안전성 테스트용).
    blank_cells: {(vendor, year)} -- 해당 셀을 아예 비워둔다 (None 값 테스트용).
    """
    headers = headers if headers is not None else BASE_HEADERS
    vendors = vendors if vendors is not None else CANONICAL_VENDORS
    value_overrides = value_overrides or {}
    blank_cells = blank_cells or set()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col, label in headers:
        ws.cell(row=HEADER_ROW, column=col, value=label)

    for i, vendor in enumerate(vendors):
        row = FIRST_VENDOR_ROW + i
        ws.cell(row=row, column=VENDOR_COL, value=vendor)
        for year, col in YEAR_COL.items():
            if (vendor, year) in blank_cells:
                continue  # leave the cell empty -> None
            value = value_overrides.get((vendor, year), vendor_value(i, year, month_variant))
            ws.cell(row=row, column=col, value=value)

    return wb


def make_valid_fixtures(out_dir) -> tuple:
    """6월/7월 정상 픽스처를 생성한다.

    6월 픽스처에는 NULL-안전성 테스트용으로 Huawei/2025를 빈 칸(None)으로,
    Honor/2025를 명시적 0으로 만들어 둔다 (v_yoy prev=NULL / prev=0 케이스).
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    jun_path = out_dir / "★(MI) Smartphone SET TAM_26년 6월 연전망.xlsx"
    jul_path = out_dir / "★(MI) Smartphone SET TAM_26년 7월 연전망.xlsx"

    build_workbook(
        "6월 연전망용",
        month_variant=6,
        blank_cells={("Huawei", 2025)},
        value_overrides={("Honor", 2025): 0.0},
    ).save(jun_path)

    build_workbook("7월 연전망용", month_variant=7).save(jul_path)

    return jun_path, jul_path


def make_jun_revised_fixture(out_dir):
    """6월과 같은 vintage(2026-06)로 재적재할 때 값이 실제로 교체되는지 보기 위한 변형본.

    MX/2026 값만 999.0으로 바꾼다 (원래 105.0).
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "★(MI) Smartphone SET TAM_26년 6월 연전망_revised.xlsx"
    build_workbook(
        "6월 연전망용",
        month_variant=6,
        blank_cells={("Huawei", 2025)},
        value_overrides={("Honor", 2025): 0.0, ("MX", 2026): 999.0},
    ).save(path)
    return path


def make_ambiguous_header_fixture(out_dir):
    """'26 라벨이 중복된 헤더 -- locator가 모호성 에러를 내야 한다."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    headers = BASE_HEADERS + [(11, "'26")]  # 중복 '26 라벨 추가 (미사용 컬럼)
    path = out_dir / "fixture_ambiguous_header.xlsx"
    build_workbook("6월 연전망용", headers=headers).save(path)
    return path


def make_no_apostrophe_year_fixture(out_dir):
    """'26 라벨을 어퍼스트로피 없는 "26"/"2026"으로 바꾼 헤더.

    어퍼스트로피가 없는 연도 형태 라벨은 매치되면 안 되므로, 2026 컬럼이 "없음"으로
    처리되어 locator가 missing-year 에러를 내야 한다.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    headers = [(col, "26" if label == "'26" else label) for col, label in BASE_HEADERS]
    headers.append((11, "2026"))  # 어퍼스트로피 없는 또 다른 변형도 같이 넣어둔다
    path = out_dir / "fixture_no_apostrophe_year.xlsx"
    build_workbook("6월 연전망용", headers=headers).save(path)
    return path


def make_missing_vendor_fixture(out_dir):
    """벤더 하나가 누락된 픽스처 (Local 자리에 MX를 중복 배치) -- 캐노니컬 벤더 누락 에러 유발."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    vendors = list(CANONICAL_VENDORS)
    vendors[-2] = vendors[0]  # "Local"(뒤에서 2번째) 자리를 "MX"로 덮어써 Local을 누락시킴
    path = out_dir / "fixture_missing_vendor.xlsx"
    build_workbook("6월 연전망용", vendors=vendors).save(path)
    return path


if __name__ == "__main__":
    target_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("fixtures")
    make_valid_fixtures(target_dir)
    print(f"Generated 2 fixture files in: {target_dir.resolve()}")
